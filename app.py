import base64, io, json, os, time
from pathlib import Path
import requests
from flask import Flask, jsonify, render_template
from openpyxl import load_workbook

app = Flask(__name__)
CACHE_SECONDS = int(os.getenv("CACHE_SECONDS", "60"))
ONEDRIVE_URL = os.getenv("ONEDRIVE_URL", "").strip()
_cache = {"items": None, "time": 0, "source": ""}

def estado(stock):
    try: n = float(stock or 0)
    except (TypeError, ValueError): n = 0
    if n >= 3: return "Disponible"
    if n >= 1: return "Ultimas unidades"
    return "No disponible"

def parse_excel(content):
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True, keep_vba=True)
    ws = wb["Inventario"]
    first = next(ws.iter_rows(values_only=True))
    headers = {str(v).strip(): i for i, v in enumerate(first) if v is not None}
    for needed in ("Marca", "Producto", "Stock Actual"):
        if needed not in headers:
            raise ValueError(f"No existe la columna {needed}")
    items=[]
    for row in ws.iter_rows(min_row=2, values_only=True):
        marca, producto = row[headers["Marca"]], row[headers["Producto"]]
        if not marca or not producto: continue
        items.append({
            "marca": str(marca).strip(),
            "producto": str(producto).strip(),
            "estado": estado(row[headers["Stock Actual"]])
        })
    return sorted(items, key=lambda x:(x["marca"].casefold(),x["producto"].casefold()))

def download_onedrive():
    # El enlace se guarda como variable privada del servidor, nunca se manda al navegador.
    candidates = [ONEDRIVE_URL]
    sep = '&' if '?' in ONEDRIVE_URL else '?'
    candidates.append(ONEDRIVE_URL + sep + 'download=1')
    last = None
    for url in candidates:
        try:
            r=requests.get(url,timeout=25,allow_redirects=True)
            r.raise_for_status()
            if r.content[:2] == b'PK': return r.content
            last=ValueError('OneDrive devolvio una pagina y no el archivo')
        except Exception as e: last=e
    raise last or RuntimeError('No se pudo descargar el Excel')

def load_items():
    now=time.time()
    if _cache["items"] is not None and now-_cache["time"]<CACHE_SECONDS:
        return _cache["items"], _cache["source"]
    try:
        if not ONEDRIVE_URL: raise RuntimeError('ONEDRIVE_URL no configurada')
        items=parse_excel(download_onedrive())
        source='OneDrive'
    except Exception:
        items=json.loads(Path(app.root_path,'inventory.json').read_text(encoding='utf-8'))
        source='copia de respaldo'
    _cache.update(items=items,time=now,source=source)
    return items,source

@app.get('/')
def home(): return render_template('index.html')

@app.get('/api/productos')
def productos():
    items,source=load_items()
    return jsonify({"productos":items,"actualizado":int(_cache['time']),"fuente":source})

@app.get('/health')
def health(): return {"ok":True}

if __name__ == '__main__':
    app.run(host='0.0.0.0',port=int(os.getenv('PORT','5000')),debug=True)
